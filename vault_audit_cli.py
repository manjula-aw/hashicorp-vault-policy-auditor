import argparse
import sys
import os
import hcl2
import json
import fnmatch
import datetime
import html
import re
import openpyxl
from openpyxl.styles import Font, PatternFill

class VaultPolicyAuditor:
    def __init__(self, input_dir, ignore_extensions=False):
        self.input_dir = input_dir
        self.ignore_extensions = ignore_extensions
        self.policies_data = {}
        self.path_matrix = {}
        self.all_concrete_paths = set()
        self.audit_issues = []
        self.processing_log = []
        self.stats = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}

    def run(self):
        self._load_files()
        self._analyze()
        return self.audit_issues

    def _load_files(self):
        if not os.path.exists(self.input_dir):
            print(f"Error: Directory not found: {self.input_dir}")
            sys.exit(1)

        print(f"Scanning directory: {self.input_dir}...")
        for root, _, files in os.walk(self.input_dir):
            for filename in files:
                if filename.startswith('.'): continue
                
                _, ext = os.path.splitext(filename)
                if self.ignore_extensions and ext != "":
                    continue
                
                filepath = os.path.join(root, filename)
                try:
                    with open(filepath, 'r') as f:
                        raw = f.read()
                    with open(filepath, 'r') as f:
                        parsed = hcl2.load(f)
                    
                    self.policies_data[filename] = {'parsed': parsed, 'raw': raw}
                    self.processing_log.append({"file": filename, "status": "SUCCESS", "msg": "Parsed OK"})
                    
                    for path_block in parsed.get('path', []):
                        for path_str, _ in path_block.items():
                            if "*" not in path_str:
                                self.all_concrete_paths.add(path_str)
                except Exception as e:
                    self.processing_log.append({"file": filename, "status": "FAILED", "msg": str(e)})
                    print(f"  [WARN] Failed to parse {filename}: {e}")

    def _analyze(self):
        # 1. Explicit Analysis
        for policy_name, data in self.policies_data.items():
            paths = data['parsed'].get('path', [])
            for path_entry in paths:
                for path_str, rules in path_entry.items():
                    caps = rules.get('capabilities', [])
                    if isinstance(caps, str): caps = [caps]
                    
                    if path_str not in self.path_matrix: self.path_matrix[path_str] = []
                    self.path_matrix[path_str].append({"policy": policy_name, "caps": caps, "via": None})
                    self._check_security(policy_name, path_str, caps)

        # 2. Wildcard Injection
        for concrete_path in self.all_concrete_paths:
            for policy_name, data in self.policies_data.items():
                paths = data['parsed'].get('path', [])
                for path_entry in paths:
                    for rule_path, rules in path_entry.items():
                        if "*" in rule_path and fnmatch.fnmatch(concrete_path, rule_path):
                            exists = any(x for x in self.path_matrix.get(concrete_path, []) 
                                         if x['policy'] == policy_name and x['via'] is None)
                            if not exists:
                                caps = rules.get('capabilities', [])
                                if isinstance(caps, str): caps = [caps]
                                self.path_matrix[concrete_path].append(
                                    {"policy": policy_name, "caps": caps, "via": rule_path}
                                )

    def _check_security(self, policy, path, caps):
        caps_lower = [c.lower() for c in caps]
        issue = None
        
        if "sudo" in caps_lower:
            issue = {"sev": "CRITICAL", "msg": "Grants 'sudo' capability", "fix": "Remove 'sudo'."}
        elif "*" in caps_lower:
            issue = {"sev": "CRITICAL", "msg": "Grants '*' capability", "fix": "Replace '*' with specific list."}
        elif path.startswith("sys/") and any(x in caps_lower for x in ["create", "update", "delete", "sudo"]):
            issue = {"sev": "HIGH", "msg": "Write access to System Backend", "fix": "Restrict to read-only."}
        elif path == "*" or path == "/*":
            issue = {"sev": "HIGH", "msg": "Root wildcard path", "fix": "Scope to specific paths."}

        if issue:
            issue.update({"pol": policy, "path": path})
            self.audit_issues.append(issue)
            self.stats[issue['sev']] += 1

    def sanitize_id(self, s):
        return re.sub(r'[^a-zA-Z0-9]', '_', s)

    def generate_html_report(self, output_file):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Build Graph
        graph_def = "graph LR\n"
        has_graph = False
        for issue in self.audit_issues:
            if issue['sev'] in ["CRITICAL", "HIGH"]:
                pol_id = self.sanitize_id(issue['pol'])
                path_id = self.sanitize_id(issue['path'])
                graph_def += f"    {pol_id}[\"{html.escape(issue['pol'])}\"] -->|Risky| {path_id}(\"{html.escape(issue['path'])}\")\n"
                has_graph = True
        
        graph_def += "    classDef policy fill:#e1f5fe,stroke:#01579b,stroke-width:2px;\n"
        graph_def += "    classDef risk fill:#ffcdd2,stroke:#b71c1c,stroke-width:2px;\n"
        if not has_graph: graph_def += "    Ok[No High Risks Detected]:::policy\n"

        # Build Table Rows
        rows_html = ""
        for i in sorted(self.audit_issues, key=lambda x: 0 if x['sev']=="CRITICAL" else 1):
            sev_class = "bg-critical" if i['sev']=="CRITICAL" else "bg-high" if i['sev']=="HIGH" else "bg-ok"
            rows_html += f"""
            <tr>
                <td><span class="badge {sev_class}">{i['sev']}</span></td>
                <td><b>{html.escape(i['pol'])}</b></td>
                <td><span class="path-mono">{html.escape(i['path'])}</span></td>
                <td>{html.escape(i['msg'])}</td>
                <td style="color:#27ae60;font-style:italic;">{html.escape(i['fix'])}</td>
            </tr>"""

        html_content = f"""<!DOCTYPE html>
        <html lang="en"><head><meta charset="UTF-8"><title>Vault Audit Report</title>
        <script type="module">import mermaid from "https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.esm.min.mjs"; mermaid.initialize({{ startOnLoad: true }});</script>
        <style>
            body {{ font-family: 'Segoe UI', sans-serif; background: #f4f6f8; color: #2c3e50; padding: 20px; }}
            .container {{ max-width: 1200px; margin: 0 auto; }}
            .card {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.05); margin-bottom: 20px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ padding: 12px; border-bottom: 1px solid #eee; text-align: left; }}
            th {{ background: #34495e; color: white; }}
            .badge {{ padding: 4px 8px; border-radius: 12px; color: white; font-weight: bold; font-size: 0.8em; }}
            .bg-critical {{ background: #e74c3c; }} .bg-high {{ background: #f39c12; }} .bg-ok {{ background: #27ae60; }}
            .path-mono {{ font-family: monospace; color: #e83e8c; background: #fdf0f5; padding: 2px 5px; border-radius: 4px; }}
            .mermaid {{ text-align: center; margin-top: 20px; }}
        </style></head><body>
        <div class="container">
            <div class="card">
                <h1>Vault Policy Audit Report</h1>
                <p>Generated: {timestamp} | Files: {len(self.policies_data)}</p>
                <p><b>CRITICAL:</b> {self.stats['CRITICAL']} | <b>HIGH:</b> {self.stats['HIGH']}</p>
            </div>
            <div class="card">
                <h2>Risk Graph</h2>
                <div class="mermaid">{graph_def}</div>
            </div>
            <div class="card">
                <h2>Security Risks</h2>
                <table><tr><th>Severity</th><th>Policy</th><th>Path</th><th>Issue</th><th>Fix</th></tr>{rows_html}</table>
            </div>
        </div></body></html>"""
        
        with open(output_file, "w", encoding="utf-8") as f:
            f.write(html_content)
        print(f"HTML Report generated: {output_file}")

    def generate_excel_report(self, output_file):
        wb = openpyxl.Workbook()
        
        # Sheet 1: Risks
        ws = wb.active
        ws.title = "Security Risks"
        ws.append(["Severity", "Policy", "Path", "Issue", "Recommendation"])
        
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for i in self.audit_issues:
            row = [i['sev'], i['pol'], i['path'], i['msg'], i['fix']]
            ws.append(row)
            if i['sev'] == "CRITICAL":
                ws.cell(row=ws.max_row, column=1).fill = PatternFill(start_color="E74C3C", fill_type="solid")

        # Sheet 2: Matrix
        ws2 = wb.create_sheet("Path Matrix")
        ws2.append(["Path", "Policy", "Via", "Capabilities", "Risk"])
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font

        for path in sorted(self.path_matrix.keys()):
            for entry in self.path_matrix[path]:
                via = entry['via'] if entry['via'] else "Direct"
                caps = ", ".join(entry['caps']).upper()
                risk = "ADMIN" if "SUDO" in caps or "*" in caps else ""
                ws2.append([path, entry['policy'], via, caps, risk])

        # Sheet 3: Log
        ws3 = wb.create_sheet("Processing Log")
        ws3.append(["File", "Status", "Message"])
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font

        for item in self.processing_log:
            ws3.append([item['file'], item['status'], item['msg']])

        wb.save(output_file)
        print(f"Excel Report generated: {output_file}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Industrialized Vault Policy Auditor (v2)")
    parser.add_argument("input_dir", help="Path to folder containing policy files")
    parser.add_argument("--html", help="Path to generate HTML report")
    parser.add_argument("--excel", help="Path to generate Excel report")
    parser.add_argument("--json", action="store_true", help="Output JSON to stdout")
    parser.add_argument("--fail-on-critical", action="store_true", help="Exit code 1 if Critical risks found")
    
    args = parser.parse_args()

    auditor = VaultPolicyAuditor(args.input_dir, ignore_extensions=True)
    issues = auditor.run()

    if args.json:
        print(json.dumps(issues, indent=2))
    
    if args.html:
        auditor.generate_html_report(args.html)
        
    if args.excel:
        auditor.generate_excel_report(args.excel)

    if args.fail_on_critical and auditor.stats["CRITICAL"] > 0:
        print(f"\n[FAILURE] Found {auditor.stats['CRITICAL']} CRITICAL issues.")
        sys.exit(1)