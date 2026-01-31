import os
import hcl2
import html
import datetime
import re
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill

class VaultAuditEngine:
    def __init__(self):
        self.policies_data = {}       
        self.path_matrix = {}         
        self.all_concrete_paths = set()
        self.audit_issues = []        
        self.processing_log = []      
        self.stats = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}

    def reset(self):
        self.__init__()

    def scan_folder(self, folder_path, extensions=None):
        """
        Scans a folder for policies.
        :param extensions: List of allowed extensions (e.g. ['.hcl', '.txt']). 
                           If None or empty, scans files with NO extension.
        """
        if not os.path.exists(folder_path):
            raise FileNotFoundError(f"Directory not found: {folder_path}")

        # Normalize extensions to ensure they have dot prefix
        valid_exts = []
        if extensions:
            valid_exts = [e if e.startswith(".") else f".{e}" for e in extensions]

        for root, _, files in os.walk(folder_path):
            for filename in files:
                if filename.startswith('.'): continue
                _, ext = os.path.splitext(filename)
                
                # FILTERING LOGIC
                if not valid_exts:
                    # Default Mode: Only scan files with NO extension
                    if ext != "": continue
                else:
                    # Explicit Mode: Only scan matching extensions
                    if ext not in valid_exts: continue
                
                filepath = os.path.join(root, filename)
                try:
                    with open(filepath, 'r') as f: raw = f.read()
                    with open(filepath, 'r') as f: parsed = hcl2.load(f)
                    
                    self.policies_data[filename] = {'parsed': parsed, 'raw': raw, 'path': filepath}
                    self.processing_log.append({"file": filename, "status": "SUCCESS", "msg": "Parsed OK"})
                    
                    for path_block in parsed.get('path', []):
                        for path_str, _ in path_block.items():
                            if "*" not in path_str and "+" not in path_str:
                                self.all_concrete_paths.add(path_str)
                except Exception as e:
                    self.processing_log.append({"file": filename, "status": "FAILED", "msg": str(e)})

    def _vault_match(self, rule_path, concrete_path):
        token_star, token_plus = "___STAR___", "___PLUS___"
        safe = rule_path.replace("*", token_star).replace("+", token_plus)
        escaped = re.escape(safe)
        regex = "^" + escaped.replace(token_star, ".*").replace(token_plus, "[^/]+") + "$"
        try: return re.match(regex, concrete_path) is not None
        except: return False

    def analyze(self):
        # 1. Direct Analysis
        for policy_name, data in self.policies_data.items():
            for path_entry in data['parsed'].get('path', []):
                for path_str, rules in path_entry.items():
                    caps = rules.get('capabilities', [])
                    if isinstance(caps, str): caps = [caps]
                    if path_str not in self.path_matrix: self.path_matrix[path_str] = []
                    self.path_matrix[path_str].append({"policy": policy_name, "caps": caps, "via": None})
                    self._check_security(policy_name, path_str, caps)

        # 2. Wildcard Injection
        for concrete_path in self.all_concrete_paths:
            for policy_name, data in self.policies_data.items():
                for path_entry in data['parsed'].get('path', []):
                    for rule_path, rules in path_entry.items():
                        if ("*" in rule_path or "+" in rule_path) and self._vault_match(rule_path, concrete_path):
                            exists = any(x for x in self.path_matrix.get(concrete_path, []) if x['policy'] == policy_name and x['via'] is None)
                            if not exists:
                                caps = rules.get('capabilities', [])
                                if isinstance(caps, str): caps = [caps]
                                self.path_matrix[concrete_path].append({"policy": policy_name, "caps": caps, "via": rule_path})

    def _check_security(self, policy, path, caps):
        caps_lower = [c.lower() for c in caps]
        issue = None
        if "sudo" in caps_lower: issue = {"sev": "CRITICAL", "msg": "Grants 'sudo' capability", "fix": "Remove 'sudo'."}
        elif "*" in caps_lower: issue = {"sev": "CRITICAL", "msg": "Grants '*' capability", "fix": "Replace '*' with explicit list."}
        elif path.startswith("sys/") and any(x in caps_lower for x in ["create", "update", "delete", "sudo"]): issue = {"sev": "HIGH", "msg": "Write access to System Backend", "fix": "Restrict to read-only."}
        elif path == "*" or path == "/*": issue = {"sev": "HIGH", "msg": "Root wildcard path", "fix": "Scope to specific paths."}
        elif "+" in path: issue = {"sev": "MEDIUM", "msg": "Uses Segment Wildcard (+)", "fix": "Verify sibling path exposure."}

        if issue:
            issue.update({"pol": policy, "path": path})
            self.audit_issues.append(issue)
            if issue['sev'] in self.stats: self.stats[issue['sev']] += 1

    def sanitize_id(self, s): return re.sub(r'[^a-zA-Z0-9]', '_', s)
    def get_risk_flag(self, caps): return "⚠ ADMIN" if ("SUDO" in caps or "*" in caps) else ""

    # --- EXPORT EXCEL ---
    def export_excel(self, file_path):
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        crit_fill = PatternFill(start_color="E74C3C", fill_type="solid")
        med_fill = PatternFill(start_color="F1C40F", fill_type="solid")

        # Sort issues
        sev_priority = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
        self.audit_issues.sort(key=lambda x: sev_priority.get(x['sev'], 99))

        # Sheet 1: Risks
        ws = wb.active; ws.title = "Security Risks"
        ws.append(["Severity", "Policy", "Path", "Issue", "Recommendation"])
        for cell in ws[1]: cell.fill, cell.font = header_fill, header_font
        for i in self.audit_issues:
            ws.append([i['sev'], i['pol'], i['path'], i['msg'], i['fix']])
            if i['sev'] == "CRITICAL": ws.cell(row=ws.max_row, column=1).fill = crit_fill
            if i['sev'] == "MEDIUM": ws.cell(row=ws.max_row, column=1).fill = med_fill

        # Sheet 2: Matrix
        ws2 = wb.create_sheet("Access Matrix")
        ws2.append(["Path", "Policy", "Via", "Capabilities", "Risk"])
        for cell in ws2[1]: cell.fill, cell.font = header_fill, header_font
        for path in sorted(self.path_matrix.keys()):
            for entry in self.path_matrix[path]:
                ws2.append([path, entry['policy'], entry['via'] or "Direct", ", ".join(entry['caps']).upper(), self.get_risk_flag(entry['caps'])])

        # Sheet 3: Inspector
        ws3 = wb.create_sheet("Policy Inspector")
        ws3.append(["Policy", "Rule Path", "Capabilities", "Matches"])
        for cell in ws3[1]: cell.fill, cell.font = header_fill, header_font
        for pol_name, data in sorted(self.policies_data.items()):
            for path_block in data['parsed'].get('path', []):
                for path_str, rules in path_block.items():
                    matches_str = ""
                    if "*" in path_str or "+" in path_str:
                         m = [x for x in self.all_concrete_paths if self._vault_match(path_str, x)]
                         if m: matches_str = ", ".join(m)
                    ws3.append([pol_name, path_str, ", ".join(rules.get('capabilities', [])).upper(), matches_str])
        
        # Sheet 4: Log
        ws4 = wb.create_sheet("Processing Log")
        ws4.append(["File", "Status", "Message"])
        for cell in ws4[1]: cell.fill, cell.font = header_fill, header_font
        for item in self.processing_log: ws4.append([item['file'], item['status'], item['msg']])
        
        wb.save(file_path)

    # --- EXPORT HTML ---
    def export_html(self, file_path):
        save_dir = os.path.dirname(file_path)
        script_dir = os.path.join(save_dir, "script")
        app_dir = os.path.dirname(os.path.abspath(__file__))
        mermaid_src = os.path.join(app_dir, "mermaid.min.js")
        
        mermaid_tag = '<script type="module">import mermaid from "https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.esm.min.mjs"; mermaid.initialize({ startOnLoad: true });</script>'
        if os.path.exists(mermaid_src):
            try:
                if not os.path.exists(script_dir): os.makedirs(script_dir)
                shutil.copy2(mermaid_src, os.path.join(script_dir, "mermaid.min.js"))
                mermaid_tag = '<script src="script/mermaid.min.js"></script>\n<script>mermaid.initialize({ startOnLoad: true });</script>'
            except: pass

        sev_priority = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
        self.audit_issues.sort(key=lambda x: sev_priority.get(x['sev'], 99))

        count_crit = self.stats['CRITICAL']
        count_high = self.stats['HIGH']
        count_files = len(self.policies_data)
        count_paths = len(self.path_matrix)
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        graph_def = "graph LR\n"
        has_graph = False
        for issue in self.audit_issues:
            if issue['sev'] in ["CRITICAL", "HIGH"]:
                graph_def += f"    {self.sanitize_id(issue['pol'])}[\"{html.escape(issue['pol'])}\"] -->|Risky| {self.sanitize_id(issue['path'])}(\"{html.escape(issue['path'])}\")\n"
                has_graph = True
        graph_def += "    classDef policy fill:#e1f5fe,stroke:#01579b,stroke-width:2px;\n    classDef risk fill:#ffcdd2,stroke:#b71c1c,stroke-width:2px;\n"
        if not has_graph: graph_def += "    Ok[No High Risks Detected]:::policy\n"

        html_content = f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><title>Hashicorp Vault Policy Auditor</title>{mermaid_tag}
        <style>
            :root{{--bg:#f4f6f8;--white:#fff;--danger:#e74c3c;--warning:#f39c12;--success:#27ae60;--primary:#2c3e50; --blue-accent:#3498db;}} 
            body{{font-family:'Segoe UI', sans-serif;background:var(--bg);color:var(--primary);padding:0;margin:0;}} 
            .navbar {{position:sticky;top:0;background:var(--primary);padding:10px 20px;z-index:1000;box-shadow:0 2px 5px rgba(0,0,0,0.2);}}
            .navbar a {{color:var(--white);text-decoration:none;margin-right:20px;font-weight:600;font-size:14px;text-transform:uppercase;}}
            .navbar a:hover {{color:var(--blue-accent);}}
            .container{{max-width:1200px;margin:20px auto;padding:0 20px;}}
            .header{{background:var(--white);padding:20px;border-radius:8px;box-shadow:0 2px 4px rgba(0,0,0,0.05);margin-bottom:20px;display:flex;justify-content:space-between;align-items:center;}}
            .dashboard{{display:grid;grid-template-columns:repeat(4, 1fr);gap:20px;margin-bottom:30px;}}
            .card{{background:var(--white);padding:20px;border-radius:8px;box-shadow:0 2px 4px rgba(0,0,0,0.05);margin-bottom:20px;}}
            .stat-card{{text-align:center;}} .stat-card h3{{margin:0;font-size:2.5em;color:var(--primary);}}
            .danger h3{{color:var(--danger);}} .warning h3{{color:var(--warning);}}
            table{{width:100%;border-collapse:collapse;margin-top:15px;}} th,td{{padding:12px;border-bottom:1px solid #eee;text-align:left;}} 
            th{{background:#34495e;color:white;}} .badge{{padding:4px 8px;border-radius:12px;color:white;font-weight:bold;font-size:0.8em;}}
            .bg-critical{{background:var(--danger);}} .bg-high{{background:var(--warning);}} .bg-medium{{background:#f1c40f;color:#333;}} .bg-low{{background:#95a5a6;}} .bg-ok{{background:var(--success);}}
            .path-mono{{font-family:monospace;color:#e83e8c;background:#fdf0f5;padding:2px 5px;border-radius:3px;}} .mermaid{{text-align:center;}}
            .text-red {{color: var(--danger); font-weight:bold;}} .text-blue {{color: var(--blue-accent); font-weight:bold;}}
        </style>
        </head><body>
        
        <div class="navbar">
            <div class="container" style="margin:0; padding:0;">
                <a href="#dashboard">Dashboard</a>
                <a href="#risks">1. Security Risks</a>
                <a href="#matrix">2. Access Matrix</a>
                <a href="#inspector">3. Policy Inspector</a>
            </div>
        </div>

        <div class="container">
            <div id="dashboard" class="header">
                <div><h1>Hashicorp Vault Policy Auditor</h1><div style="color:#777">{timestamp}</div></div>
                <div><span class="badge bg-ok">v24.0</span></div>
            </div>

            <div class="dashboard">
                <div class="card stat-card"><h3>{count_files}</h3><p>Policies Scanned</p></div>
                <div class="card stat-card danger"><h3>{count_crit}</h3><p>Critical Risks</p></div>
                <div class="card stat-card warning"><h3>{count_high}</h3><p>High Risks</p></div>
                <div class="card stat-card"><h3>{count_paths}</h3><p>Unique Paths</p></div>
            </div>

            <div class="card"><h2>Risk Visualization</h2><div class="mermaid">{graph_def}</div></div>
            
            <div id="risks" class="card"><h2>1. Security Risks</h2><table><thead><tr><th>Severity</th><th>Policy</th><th>Path</th><th>Issue</th><th>Fix</th></tr></thead><tbody>"""
        
        if not self.audit_issues: html_content += "<tr><td colspan='5' style='text-align:center;color:green'>✅ No obvious security risks detected.</td></tr>"
        for i in self.audit_issues:
            sev_class = f"bg-{i['sev'].lower()}"
            sev_text_class = "text-red" if i['sev'] == "CRITICAL" else ""
            html_content += f"<tr><td><span class='badge {sev_class}'>{i['sev']}</span></td><td><b>{html.escape(i['pol'])}</b></td><td><span class='path-mono'>{html.escape(i['path'])}</span></td><td class='{sev_text_class}'>{html.escape(i['msg'])}</td><td>{html.escape(i['fix'])}</td></tr>"
        html_content += "</tbody></table></div>"

        html_content += """<div id="matrix" class="card"><h2>2. Access Matrix</h2><table><thead><tr><th>Path</th><th>Accessible By (Policy)</th><th>Capabilities</th></tr></thead><tbody>"""
        for path in sorted(self.path_matrix.keys()):
            first = True
            for e in self.path_matrix[path]:
                p_cell = f"<td rowspan='{len(self.path_matrix[path])}' style='border-right:1px solid #eee'><span class='path-mono'>{html.escape(path)}</span></td>" if first else ""
                via_txt = f"<br><small class='text-blue'>via {html.escape(e['via'])}</small>" if e['via'] else ""
                html_content += f"<tr>{p_cell}<td><b>{html.escape(e['policy'])}</b>{via_txt}</td><td>{', '.join(e['caps']).upper()}</td></tr>"
                first = False
        html_content += "</tbody></table></div>"

        html_content += """<div id="inspector" class="card"><h2>3. Policy Inspector</h2><table><thead><tr><th>Policy</th><th>Path</th><th>Matches</th></tr></thead><tbody>"""
        for pol, data in sorted(self.policies_data.items()):
            paths = []
            for pb in data['parsed'].get('path', []):
                for p_str, r in pb.items():
                    m_html = ""
                    if "*" in p_str or "+" in p_str:
                        m = [x for x in self.all_concrete_paths if self._vault_match(p_str, x)]
                        if m: m_html = "<br><small class='text-blue'>↳ " + ", ".join(m) + "</small>"
                    paths.append((p_str, ", ".join(r.get('capabilities', [])).upper(), m_html))
            
            if not paths: html_content += f"<tr><td><b>{html.escape(pol)}</b></td><td colspan='2'><i>No paths</i></td></tr>"
            else:
                for idx, (p, c, m) in enumerate(paths):
                    pol_cell = f"<td rowspan='{len(paths)}' style='border-right:1px solid #eee;vertical-align:top'><b>{html.escape(pol)}</b></td>" if idx == 0 else ""
                    html_content += f"<tr>{pol_cell}<td><span class='path-mono'>{html.escape(p)}</span><br><small>{c}</small></td><td>{m}</td></tr>"

        html_content += """</tbody></table></div>
        
        <div class="card"><h2>Processing Log</h2><table><thead><tr><th>File</th><th>Status</th><th>Details</th></tr></thead><tbody>"""
        for log in self.processing_log:
            st = "bg-ok" if log['status'] == "SUCCESS" else "bg-critical"
            html_content += f"<tr><td>{html.escape(log['file'])}</td><td><span class='badge {st}'>{log['status']}</span></td><td>{html.escape(log['msg'])}</td></tr>"
        html_content += "</tbody></table></div></div></body></html>"

        with open(file_path, "w", encoding="utf-8") as f: f.write(html_content)
