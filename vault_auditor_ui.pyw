import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import hcl2
import os
import fnmatch
import html
import datetime
import re
import openpyxl
import shutil  # Added for file copying
from openpyxl.styles import Font, PatternFill

class VaultAuditTool:
    def __init__(self, root):
        self.root = root
        self.root.title("Vault Policy Auditor (GitLab Ready)")
        self.root.geometry("1500x900")

        # Data Storage
        self.selected_files = [] 
        self.policies_data = {} 
        self.path_matrix = {}   
        self.all_concrete_paths = set() 
        self.audit_issues = []  
        self.processing_log = []

        self.var_no_ext = tk.BooleanVar(value=True) 

        # --- UI LAYOUT ---
        top_frame = tk.Frame(root, padx=10, pady=10, bg="#f0f0f0", relief="raised", borderwidth=1)
        top_frame.pack(fill=tk.X, padx=5, pady=5)

        # Row 1: Selection
        btn_frame = tk.Frame(top_frame, bg="#f0f0f0")
        btn_frame.pack(side=tk.TOP, fill=tk.X, pady=5)

        tk.Label(btn_frame, text="Step 1: Select Input", bg="#f0f0f0", font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        self.btn_folder = tk.Button(btn_frame, text="Browse Folder", command=self.browse_folder, bg="#e1e1e1")
        self.btn_folder.pack(side=tk.LEFT, padx=10)
        
        self.lbl_status = tk.Label(btn_frame, text="No folder selected", bg="#f0f0f0", fg="blue", font=("Arial", 9, "italic"))
        self.lbl_status.pack(side=tk.LEFT, padx=10)

        # Row 2: Filtering & Actions
        filter_frame = tk.Frame(top_frame, bg="#f0f0f0")
        filter_frame.pack(side=tk.TOP, fill=tk.X, pady=5)

        tk.Label(filter_frame, text="Step 2: Filter Rules:", bg="#f0f0f0", font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        chk_no_ext = tk.Checkbutton(filter_frame, text="Scan files with NO extension only", 
                                    variable=self.var_no_ext, bg="#f0f0f0", command=self.toggle_entry)
        chk_no_ext.pack(side=tk.LEFT, padx=10)

        self.lbl_ext = tk.Label(filter_frame, text="OR match extensions:", bg="#f0f0f0", fg="gray")
        self.lbl_ext.pack(side=tk.LEFT, padx=5)
        
        self.ent_extensions = tk.Entry(filter_frame, width=15, state="disabled")
        self.ent_extensions.insert(0, ".hcl, .txt")
        self.ent_extensions.pack(side=tk.LEFT)

        btn_run = tk.Button(filter_frame, text="RUN AUDIT", command=self.run_audit, bg="#4CAF50", fg="white", font=("Arial", 10, "bold"))
        btn_run.pack(side=tk.RIGHT, padx=10)

        # Export Buttons
        self.btn_export_html = tk.Button(filter_frame, text="Export HTML", command=self.export_html, bg="#008CBA", fg="white", font=("Arial", 10, "bold"), state="disabled")
        self.btn_export_html.pack(side=tk.RIGHT, padx=5)

        self.btn_export_excel = tk.Button(filter_frame, text="Export Excel", command=self.export_excel, bg="#217346", fg="white", font=("Arial", 10, "bold"), state="disabled")
        self.btn_export_excel.pack(side=tk.RIGHT, padx=5)

        # --- TABS ---
        self.tabs = ttk.Notebook(root)
        self.tabs.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Tab 1: Risks
        self.tab_risks = tk.Frame(self.tabs)
        self.tabs.add(self.tab_risks, text="1. Security Risks")
        self.tree_risks = self.create_tree(self.tab_risks, ["Severity", "Policy File", "Path", "Issue", "Recommendation"])

        # Tab 2: Matrix
        self.tab_matrix = tk.Frame(self.tabs)
        self.tabs.add(self.tab_matrix, text="2. Path Matrix")
        tb_matrix = tk.Frame(self.tab_matrix)
        tb_matrix.pack(fill=tk.X, padx=5, pady=2)
        tk.Button(tb_matrix, text="+ Expand All", command=lambda: self.expand_all(self.tree_matrix), font=("Arial", 8)).pack(side=tk.LEFT)
        tk.Button(tb_matrix, text="- Collapse All", command=lambda: self.collapse_all(self.tree_matrix), font=("Arial", 8)).pack(side=tk.LEFT, padx=5)
        self.tree_matrix = self.create_tree(self.tab_matrix, ["Capabilities", "Risk"], tree_col="#0", tree_head="Path / Policy")

        # Tab 3: Inspector
        self.tab_policies = tk.Frame(self.tabs)
        self.tabs.add(self.tab_policies, text="3. Policy Inspector")
        tb_policies = tk.Frame(self.tab_policies)
        tb_policies.pack(fill=tk.X, padx=5, pady=2)
        tk.Button(tb_policies, text="+ Expand All", command=lambda: self.expand_all(self.tree_policies), font=("Arial", 8)).pack(side=tk.LEFT)
        tk.Button(tb_policies, text="- Collapse All", command=lambda: self.collapse_all(self.tree_policies), font=("Arial", 8)).pack(side=tk.LEFT, padx=5)
        self.tree_policies = self.create_tree(self.tab_policies, ["Capabilities", "Info"], tree_col="#0", tree_head="Policy / Paths")

        # Context Menu
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="View Raw Policy", command=self.view_raw_policy)
        self.tree_policies.bind("<Button-3>", self.show_context_menu) 
        self.tree_policies.bind("<Button-2>", self.show_context_menu) 

        self.setup_styles()
        self.toggle_entry()

    def create_tree(self, parent, columns, tree_col=None, tree_head=None):
        cols = tuple(columns)
        tree = ttk.Treeview(parent, columns=cols, selectmode="browse")
        if tree_col:
            tree.heading("#0", text=tree_head)
            tree.column("#0", width=450)
        else:
            tree["show"] = "headings"
        for col in columns:
            tree.heading(col, text=col)
            width = 80 if col == "Severity" else 300
            if col == "Capabilities": width = 200
            if col == "Risk" or col == "Info": width = 150
            if col == "Recommendation": width = 400
            tree.column(col, width=width)
        sb = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(fill=tk.BOTH, expand=True)
        return tree

    def setup_styles(self):
        for tree in [self.tree_risks, self.tree_matrix, self.tree_policies]:
            tree.tag_configure("CRITICAL", foreground="white", background="#D32F2F")
            tree.tag_configure("HIGH", foreground="black", background="#FF9800")
            tree.tag_configure("MEDIUM", foreground="black", background="#FFEB3B")
            tree.tag_configure("LOW", foreground="black", background="white")
            tree.tag_configure("ERROR", foreground="white", background="gray")
        self.tree_policies.tag_configure("WILDCARD", foreground="blue")
        self.tree_policies.tag_configure("MATCH", foreground="gray")
        self.tree_matrix.tag_configure("IMPLICIT", foreground="gray")

    def expand_all(self, tree):
        def _expand_recursive(item):
            tree.item(item, open=True)
            for child in tree.get_children(item): _expand_recursive(child)
        for item in tree.get_children(): _expand_recursive(item)

    def collapse_all(self, tree):
        def _collapse_recursive(item):
            tree.item(item, open=False)
            for child in tree.get_children(item): _collapse_recursive(child)
        for item in tree.get_children(): _collapse_recursive(item)

    def toggle_entry(self):
        if self.var_no_ext.get():
            self.ent_extensions.config(state="disabled")
            self.lbl_ext.config(fg="gray")
        else:
            self.ent_extensions.config(state="normal")
            self.lbl_ext.config(fg="black")

    def browse_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.selected_files = ["FOLDER:" + folder]
            self.lbl_status.config(text=f"Selected Folder: {folder}")

    def get_files_to_scan(self):
        final_list = []
        if not self.selected_files: return []
        if self.selected_files[0].startswith("FOLDER:"):
            folder_path = self.selected_files[0].replace("FOLDER:", "")
            try:
                all_items = os.listdir(folder_path)
                for item in all_items:
                    full_path = os.path.join(folder_path, item)
                    if os.path.isfile(full_path) and not item.startswith('.'):
                        name, ext = os.path.splitext(item)
                        if self.var_no_ext.get():
                            if ext == "": final_list.append(full_path)
                        else:
                            ext_str = self.ent_extensions.get().lower()
                            allowed_exts = [x.strip() for x in ext_str.split(",") if x.strip()]
                            if not allowed_exts: final_list.append(full_path)
                            elif any(item.lower().endswith(e) for e in allowed_exts):
                                final_list.append(full_path)
            except Exception as e:
                messagebox.showerror("Error", str(e))
                return []
        else:
            final_list = self.selected_files
        return final_list

    def run_audit(self):
        files_to_scan = self.get_files_to_scan()
        if not files_to_scan:
            messagebox.showwarning("Warning", "No matching files found.")
            return
        
        # Reset Data
        self.policies_data = {}
        self.path_matrix = {}
        self.audit_issues = []
        self.all_concrete_paths = set()
        self.processing_log = []
        for tree in [self.tree_risks, self.tree_matrix, self.tree_policies]:
            for item in tree.get_children(): tree.delete(item)

        parsed_count = 0
        for filepath in files_to_scan:
            filename = os.path.basename(filepath)
            try:
                with open(filepath, 'r') as f: raw_content = f.read()
                with open(filepath, 'r') as f: parsed_dict = hcl2.load(f)
                
                self.policies_data[filename] = {'parsed': parsed_dict, 'raw': raw_content, 'path': filepath}
                parsed_count += 1
                self.processing_log.append({"file": filename, "status": "SUCCESS", "msg": "Parsed OK"})
                
                for path_block in parsed_dict.get('path', []):
                    for path_str, _ in path_block.items():
                        if "*" not in path_str: self.all_concrete_paths.add(path_str)
            except Exception as e:
                self.processing_log.append({"file": filename, "status": "FAILED", "msg": str(e)})
                self.tree_risks.insert("", "end", values=("ERROR", filename, "N/A", f"Parse Failed: {str(e)}", "Check file syntax"), tags=("ERROR",))

        self.analyze_policies()
        self.populate_ui()
        self.btn_export_html.config(state="normal")
        self.btn_export_excel.config(state="normal")
        messagebox.showinfo("Done", f"Scanned {len(files_to_scan)} files.\nParsed {parsed_count} successfully.")

    def analyze_policies(self):
        for policy_name, data_obj in self.policies_data.items():
            content = data_obj['parsed']
            paths = content.get('path', [])
            for path_entry in paths:
                for path_str, rules in path_entry.items():
                    caps = rules.get('capabilities', [])
                    if isinstance(caps, str): caps = [caps]
                    if path_str not in self.path_matrix: self.path_matrix[path_str] = []
                    self.path_matrix[path_str].append({"policy": policy_name, "caps": caps, "via": None})
                    self.check_security(policy_name, path_str, caps)
        
        for concrete_path in self.all_concrete_paths:
            for policy_name, data_obj in self.policies_data.items():
                content = data_obj['parsed']
                already_explicit = False
                for existing in self.path_matrix.get(concrete_path, []):
                    if existing['policy'] == policy_name and existing['via'] is None:
                        already_explicit = True
                        break
                if already_explicit: continue 

                for path_entry in content.get('path', []):
                    for rule_path, rules in path_entry.items():
                        if "*" in rule_path and fnmatch.fnmatch(concrete_path, rule_path):
                            caps = rules.get('capabilities', [])
                            if isinstance(caps, str): caps = [caps]
                            self.path_matrix[concrete_path].append({"policy": policy_name, "caps": caps, "via": rule_path})

    def check_security(self, policy, path, caps):
        caps_lower = [c.lower() for c in caps]
        if "sudo" in caps_lower: 
            self.audit_issues.append({"sev": "CRITICAL", "pol": policy, "path": path, "msg": "Grants 'sudo' capability", "fix": "Remove 'sudo' from capabilities."})
        if "*" in caps_lower: 
            self.audit_issues.append({"sev": "CRITICAL", "pol": policy, "path": path, "msg": "Grants '*' capability", "fix": "Replace '*' with specific list [\"read\", etc]."})
        if path.startswith("sys/") and any(x in caps_lower for x in ["create", "update", "delete", "sudo"]):
             self.audit_issues.append({"sev": "HIGH", "pol": policy, "path": path, "msg": "Write access to System Backend", "fix": "Restrict to read-only or specific sub-paths."})
        if path == "*" or path == "/*": 
            self.audit_issues.append({"sev": "HIGH", "pol": policy, "path": path, "msg": "Root wildcard path", "fix": "Scope this policy to specific paths."})

    def get_risk_flag(self, caps):
        caps_str = ", ".join(caps).upper()
        if "SUDO" in caps_str or "*" in caps_str: return "⚠ ADMIN"
        return ""

    def populate_ui(self):
        sev_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "ERROR": 4}
        self.audit_issues.sort(key=lambda x: sev_order.get(x["sev"], 99))
        for i in self.audit_issues:
            self.tree_risks.insert("", "end", values=(i["sev"], i["pol"], i["path"], i["msg"], i["fix"]), tags=(i["sev"],))

        for path in sorted(self.path_matrix.keys()):
            node = self.tree_matrix.insert("", "end", text=path, open=False)
            entries = self.path_matrix[path]
            entries.sort(key=lambda x: (x['via'] is not None, x['policy']))
            for entry in entries:
                caps_str = ", ".join(entry["caps"]).upper()
                risk = self.get_risk_flag(entry["caps"])
                display = entry["policy"] + (f" (via {entry['via']})" if entry['via'] else "")
                tags = ("IMPLICIT",) if entry['via'] else ()
                self.tree_matrix.insert(node, "end", text=display, values=(caps_str, risk), tags=tags)

        for policy in sorted(self.policies_data.keys()):
            p_node = self.tree_policies.insert("", "end", text=policy, open=False)
            content = self.policies_data[policy]['parsed']
            for path_block in content.get('path', []):
                for path_str, rules in path_block.items():
                    caps = rules.get('capabilities', [])
                    if isinstance(caps, str): caps = [caps]
                    item_id = self.tree_policies.insert(p_node, "end", text=path_str, values=(", ".join(caps).upper(), self.get_risk_flag(caps)))
                    if "*" in path_str:
                        matches = [m for m in self.all_concrete_paths if fnmatch.fnmatch(m, path_str)]
                        if matches:
                            self.tree_policies.item(item_id, tags=("WILDCARD",))
                            for m in matches: self.tree_policies.insert(item_id, "end", text=f"↳ Matches: {m}", values=("(Inherited)", "See Parent"), tags=("MATCH",))

    def show_context_menu(self, event):
        item_id = self.tree_policies.identify_row(event.y)
        if item_id:
            self.tree_policies.selection_set(item_id)
            if self.tree_policies.item(item_id, "text") in self.policies_data:
                self.context_menu.post(event.x_root, event.y_root)

    def view_raw_policy(self):
        sel = self.tree_policies.selection()
        if not sel: return
        name = self.tree_policies.item(sel[0], "text")
        if name in self.policies_data:
            top = tk.Toplevel(self.root)
            top.title(f"Viewing: {name}")
            top.geometry("600x600")
            txt = tk.Text(top, wrap="word", padx=10, pady=10)
            txt.insert("1.0", self.policies_data[name]['raw'])
            txt.config(state="disabled")
            txt.pack(fill=tk.BOTH, expand=True)

    def sanitize_id(self, s):
        return re.sub(r'[^a-zA-Z0-9]', '_', s)

    # --- EXCEL EXPORT ---
    def export_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile="vault-audit-report.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not file_path: return

        try:
            wb = openpyxl.Workbook()
            ws_risks = wb.active
            ws_risks.title = "Security Risks"
            ws_risks.append(["Severity", "Policy File", "Path", "Issue", "Recommendation"])
            
            header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws_risks[1]: cell.fill, cell.font = header_fill, header_font

            for issue in self.audit_issues:
                row = [issue['sev'], issue['pol'], issue['path'], issue['msg'], issue['fix']]
                ws_risks.append(row)
                if issue['sev'] == "CRITICAL":
                    ws_risks.cell(row=ws_risks.max_row, column=1).fill = PatternFill(start_color="E74C3C", fill_type="solid")

            ws_matrix = wb.create_sheet("Access Matrix")
            ws_matrix.append(["Path", "Policy", "Via Wildcard", "Capabilities", "Risk Flag"])
            for cell in ws_matrix[1]: cell.fill, cell.font = header_fill, header_font

            for path in sorted(self.path_matrix.keys()):
                entries = self.path_matrix[path]
                entries.sort(key=lambda x: (x['via'] is not None, x['policy']))
                for entry in entries:
                    via_text = entry['via'] if entry['via'] else "Direct"
                    ws_matrix.append([path, entry['policy'], via_text, ", ".join(entry["caps"]).upper(), self.get_risk_flag(entry["caps"])])

            wb.save(file_path)
            messagebox.showinfo("Export Success", f"Excel report saved to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

    def export_html(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".html",
            initialfile="vault-policy-audit-report.html",
            filetypes=[("HTML files", "*.html")]
        )
        if not file_path: return

        # --- LOGIC TO COPY MERMAID FILE ---
        # 1. Determine destination directory
        save_dir = os.path.dirname(file_path)
        script_dir = os.path.join(save_dir, "script")
        
        # 2. Determine source directory (where this python script is running)
        app_dir = os.path.dirname(os.path.abspath(__file__))
        mermaid_src = os.path.join(app_dir, "mermaid.min.js")
        mermaid_dest = os.path.join(script_dir, "mermaid.min.js")
        
        mermaid_script_tag = ""
        
        try:
            # Check if source file exists
            if os.path.exists(mermaid_src):
                # Create 'script' folder if missing
                if not os.path.exists(script_dir):
                    os.makedirs(script_dir)
                
                # Copy file
                shutil.copy2(mermaid_src, mermaid_dest)
                print(f"Copied mermaid.min.js to {mermaid_dest}")
                
                # Use Relative Path in HTML
                mermaid_script_tag = '<script src="script/mermaid.min.js"></script>'
            else:
                print("Local mermaid.min.js not found. Falling back to CDN.")
                mermaid_script_tag = '<script type="module">import mermaid from "https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.esm.min.mjs"; mermaid.initialize({ startOnLoad: true });</script>'

        except Exception as e:
            print(f"Error copying mermaid file: {e}")
            # Fallback
            mermaid_script_tag = '<script type="module">import mermaid from "https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.esm.min.mjs"; mermaid.initialize({ startOnLoad: true });</script>'

        # Add Initialization if using local file
        if "src=" in mermaid_script_tag:
            mermaid_script_tag += "\n<script>mermaid.initialize({ startOnLoad: true });</script>"

        # Stats
        count_crit = len([x for x in self.audit_issues if x['sev'] == "CRITICAL"])
        count_high = len([x for x in self.audit_issues if x['sev'] == "HIGH"])
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Generate Graph
        graph_def = "graph LR\n"
        has_graph_data = False
        for issue in self.audit_issues:
            if issue['sev'] in ["CRITICAL", "HIGH"]:
                pol_id = self.sanitize_id(issue['pol'])
                path_id = self.sanitize_id(issue['path'])
                graph_def += f"    {pol_id}[\"{html.escape(issue['pol'])}\"] -->|Risky| {path_id}(\"{html.escape(issue['path'])}\")\n"
                has_graph_data = True
        
        graph_def += "    classDef policy fill:#e1f5fe,stroke:#01579b,stroke-width:2px;\n"
        graph_def += "    classDef risk fill:#ffcdd2,stroke:#b71c1c,stroke-width:2px;\n"
        if not has_graph_data: graph_def += "    Ok[No High Risks Detected]:::policy\n"

        html_content = f"""<!DOCTYPE html>
        <html lang="en">
        <head>
        <meta charset="UTF-8">
        <title>Vault Policy Audit Report</title>
        {mermaid_script_tag}
        <style>
            :root {{ --primary: #2c3e50; --secondary: #34495e; --bg: #f4f6f8; --white: #ffffff; --danger: #e74c3c; --warning: #f39c12; }}
            body {{ font-family: 'Segoe UI', sans-serif; background: var(--bg); color: var(--primary); padding: 20px; }}
            .container {{ max-width: 1200px; margin: 0 auto; }}
            .card {{ background: var(--white); padding: 20px; border-radius: 8px; margin-bottom: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.05); }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th, td {{ padding: 12px; border-bottom: 1px solid #eee; text-align: left; }}
            th {{ background: var(--secondary); color: white; }}
            .mermaid {{ text-align: center; }}
            .badge {{ padding: 4px 8px; border-radius: 12px; color: white; font-weight: bold; font-size: 0.8em; }}
            .bg-critical {{ background: var(--danger); }} .bg-high {{ background: var(--warning); }}
        </style>
        </head>
        <body>
        <div class="container">
            <div class="card">
                <h1>Vault Policy Audit Report</h1>
                <p>Generated: {timestamp}</p>
            </div>
            <div class="card">
                <h2>Risk Graph</h2>
                <div class="mermaid">{graph_def}</div>
            </div>
            <div class="card">
                <h2>Security Risks</h2>
                <table><thead><tr><th>Severity</th><th>Policy</th><th>Path</th><th>Issue</th><th>Fix</th></tr></thead><tbody>
        """
        for issue in self.audit_issues:
            sev_class = f"bg-{issue['sev'].lower()}"
            html_content += f"<tr><td><span class='badge {sev_class}'>{issue['sev']}</span></td><td>{html.escape(issue['pol'])}</td><td>{html.escape(issue['path'])}</td><td>{html.escape(issue['msg'])}</td><td>{html.escape(issue['fix'])}</td></tr>"
        
        html_content += "</tbody></table></div></div></body></html>"

        try:
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(html_content)
            messagebox.showinfo("Export Success", f"Report saved to:\n{file_path}\n\nExternal script copied to:\n{mermaid_dest}")
        except Exception as e:
            messagebox.showerror("Export Failed", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = VaultAuditTool(root)
    root.mainloop()